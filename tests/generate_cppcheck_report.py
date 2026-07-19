#!/usr/bin/env python3
import os
import sys
import xml.etree.ElementTree as ET

def generate_cppcheck_report(xml_path, xlsx_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if not os.path.exists(xml_path):
        print(f"Cppcheck XML file not found at: {xml_path}")
        return False

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except Exception as e:
        print(f"Error parsing Cppcheck XML {xml_path}: {e}")
        return False

    errors_list = []
    # Find all error elements
    for error in root.findall(".//error"):
        error_id = error.attrib.get("id", "")
        severity = error.attrib.get("severity", "")
        msg = error.attrib.get("msg", "")
        verbose = error.attrib.get("verbose", "")
        cwe = error.attrib.get("cwe", "")
        
        locations = error.findall("location")
        if not locations:
            errors_list.append({
                "file": "N/A",
                "line": "N/A",
                "column": "N/A",
                "severity": severity.capitalize(),
                "id": error_id,
                "msg": msg,
                "verbose": verbose,
                "cwe": cwe
            })
        else:
            for loc in locations:
                errors_list.append({
                    "file": loc.attrib.get("file", ""),
                    "line": loc.attrib.get("line", ""),
                    "column": loc.attrib.get("column", ""),
                    "severity": severity.capitalize(),
                    "id": error_id,
                    "msg": msg,
                    "verbose": verbose,
                    "cwe": cwe
                })

    if has_openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cppcheck Report"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )

        # Fills for severities
        fill_error = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
        font_error = Font(name="Calibri", size=11, color="C00000", bold=True)
        fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft yellow
        font_warning = Font(name="Calibri", size=11, color="7F6000", bold=True)
        fill_style = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
        font_style = Font(name="Calibri", size=11, color="375623")

        # Title
        ws.cell(row=1, column=1, value="Cppcheck Static Analysis Report").font = title_font

        # Command Invocation
        command_font = Font(name="Calibri", size=9, italic=True, color="555555")
        ws.cell(row=2, column=1, value="Command: cppcheck --enable=warning,style,performance,portability --inline-suppr --std=c++17 --error-exitcode=1 --xml --xml-version=2 src/ tests/").font = command_font

        # Summary Metrics
        ws.cell(row=3, column=1, value="Metric").font = bold_font
        ws.cell(row=3, column=1).fill = sub_header_fill
        ws.cell(row=3, column=1).border = thin_border
        ws.cell(row=3, column=2, value="Count").font = bold_font
        ws.cell(row=3, column=2).fill = sub_header_fill
        ws.cell(row=3, column=2).border = thin_border

        total_issues = len(errors_list)
        errors_count = sum(1 for e in errors_list if e["severity"].lower() == "error")
        warnings_count = sum(1 for e in errors_list if e["severity"].lower() == "warning")
        style_count = total_issues - errors_count - warnings_count

        metrics = [
            ("Total Issues", total_issues),
            ("Errors", errors_count),
            ("Warnings", warnings_count),
            ("Style/Perf/Portability/Info", style_count)
        ]

        for idx, (m, v) in enumerate(metrics, start=4):
            ws.cell(row=idx, column=1, value=m).font = regular_font
            ws.cell(row=idx, column=1).border = thin_border
            val_cell = ws.cell(row=idx, column=2, value=v)
            val_cell.font = bold_font
            val_cell.border = thin_border
            if m == "Errors" and errors_count > 0:
                val_cell.fill = fill_error
                val_cell.font = font_error
            elif m == "Warnings" and warnings_count > 0:
                val_cell.fill = fill_warning
                val_cell.font = font_warning

        # Headers
        headers = ["File", "Line", "Column", "Severity", "Error ID", "Message", "CWE", "Details"]
        start_row = 10
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Rows
        for row_idx, err in enumerate(errors_list, start=start_row + 1):
            c1 = ws.cell(row=row_idx, column=1, value=err["file"])
            c2 = ws.cell(row=row_idx, column=2, value=err["line"])
            c3 = ws.cell(row=row_idx, column=3, value=err["column"])
            c4 = ws.cell(row=row_idx, column=4, value=err["severity"])
            c5 = ws.cell(row=row_idx, column=5, value=err["id"])
            c6 = ws.cell(row=row_idx, column=6, value=err["msg"])
            c7 = ws.cell(row=row_idx, column=7, value=err["cwe"])
            c8 = ws.cell(row=row_idx, column=8, value=err["verbose"])

            for c in [c1, c2, c3, c4, c5, c6, c7, c8]:
                c.font = regular_font
                c.border = thin_border

            try:
                if err["line"] != "N/A":
                    c2.value = int(err["line"])
                if err["column"] != "N/A":
                    c3.value = int(err["column"])
            except ValueError:
                pass

            c2.alignment = Alignment(horizontal="right")
            c3.alignment = Alignment(horizontal="right")
            c4.alignment = Alignment(horizontal="center")
            c7.alignment = Alignment(horizontal="center")

            # Style severity cell
            sev = err["severity"].lower()
            if sev == "error":
                c4.fill = fill_error
                c4.font = font_error
            elif sev == "warning":
                c4.fill = fill_warning
                c4.font = font_warning
            else:
                c4.fill = fill_style
                c4.font = font_style

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < start_row:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Apply specific widths for long text columns
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['H'].width = 50

        wb.save(xlsx_path)
        print(f"Cppcheck Excel report generated successfully at: {xlsx_path}")
        return True
    else:
        print("openpyxl is not installed. Generating HTML-based Excel report instead...")
        with open(xlsx_path, "w", encoding="utf-8") as f:
            f.write("""<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel" xmlns="http://www.w3.org/TR/REC-html40">
<head>
<meta http-equiv="content-type" content="text/html; charset=utf-8">
<style>
  body { font-family: Calibri, sans-serif; }
  table { border-collapse: collapse; margin-bottom: 30px; }
  th { background-color: #1F497D; color: white; font-weight: bold; }
  td, th { border: 1px solid #BFBFBF; text-align: left; padding: 8px; }
  .metric-header { background-color: #DCE6F1; font-weight: bold; }
  .severity-error { background-color: #FCE4D6; color: #C00000; font-weight: bold; text-align: center; }
  .severity-warning { background-color: #FFF2CC; color: #7F6000; font-weight: bold; text-align: center; }
  .severity-style { background-color: #E2EFDA; color: #375623; text-align: center; }
  .title { font-size: 16px; font-weight: bold; color: #1F497D; margin-bottom: 10px; }
</style>
</head>
<body>
<div class="title">Cppcheck Static Analysis Report</div>
<div style="font-size: 11px; font-style: italic; color: #555555; margin-bottom: 15px;">Command: cppcheck --enable=warning,style,performance,portability --inline-suppr --std=c++17 --error-exitcode=1 --xml --xml-version=2 src/ tests/</div>
""")
            total_issues = len(errors_list)
            errors_count = sum(1 for e in errors_list if e["severity"].lower() == "error")
            warnings_count = sum(1 for e in errors_list if e["severity"].lower() == "warning")
            style_count = total_issues - errors_count - warnings_count

            f.write('<table>\n')
            f.write(f'  <tr><td class="metric-header">Total Issues</td><td><b>{total_issues}</b></td></tr>\n')
            f.write(f'  <tr><td class="metric-header">Errors</td><td><b>{errors_count}</b></td></tr>\n')
            f.write(f'  <tr><td class="metric-header">Warnings</td><td><b>{warnings_count}</b></td></tr>\n')
            f.write(f'  <tr><td class="metric-header">Style/Perf/Portability/Info</td><td><b>{style_count}</b></td></tr>\n')
            f.write('</table>\n')

            f.write('<table>\n')
            f.write('  <tr><th>File</th><th>Line</th><th>Column</th><th>Severity</th><th>Error ID</th><th>Message</th><th>CWE</th><th>Details</th></tr>\n')
            for err in errors_list:
                sev = err["severity"].lower()
                if sev == "error":
                    sev_class = "severity-error"
                elif sev == "warning":
                    sev_class = "severity-warning"
                else:
                    sev_class = "severity-style"

                safe_msg = err["msg"].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                safe_verbose = err["verbose"].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                f.write(f'  <tr><td>{err["file"]}</td><td style="text-align: right;">{err["line"]}</td><td style="text-align: right;">{err["column"]}</td><td class="{sev_class}">{err["severity"]}</td><td>{err["id"]}</td><td>{safe_msg}</td><td>{err["cwe"]}</td><td>{safe_verbose}</td></tr>\n')
            f.write('</table>\n</body>\n</html>\n')
        print(f"HTML Cppcheck report generated successfully at: {xlsx_path}")
        return True

if __name__ == "__main__":
    xml_p = "cppcheck_report.xml"
    xlsx_p = "cppcheck_report.xlsx"
    if len(sys.argv) > 1:
        xml_p = sys.argv[1]
    if len(sys.argv) > 2:
        xlsx_p = sys.argv[2]
    generate_cppcheck_report(xml_p, xlsx_p)
