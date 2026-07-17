#!/usr/bin/env python3
import os
import glob
import xml.etree.ElementTree as ET
import sys

def parse_ctest_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except Exception as e:
        print(f"Error parsing {xml_path}: {e}")
        return []

    tests = []
    # Find all Test elements under Testing
    for test_elem in root.findall(".//Testing/Test"):
        name = test_elem.find("Name").text if test_elem.find("Name") is not None else "Unknown"
        status = test_elem.attrib.get("Status", "unknown")
        
        duration = "0.0"
        for measurement in test_elem.findall(".//NamedMeasurement"):
            if measurement.attrib.get("name") == "Execution Time":
                val_elem = measurement.find("Value")
                if val_elem is not None:
                    duration = val_elem.text
                    break
        
        # Get console output if any
        output = ""
        meas_elem = test_elem.find(".//Measurement")
        if meas_elem is not None:
            val_elem = meas_elem.find("Value")
            if val_elem is not None and val_elem.text:
                raw_text = val_elem.text
                if val_elem.attrib.get("encoding") == "base64":
                    try:
                        import base64
                        decoded_data = base64.b64decode(raw_text)
                        
                        # Try to decompress if compression attribute is present or data looks compressed
                        compression = val_elem.attrib.get("compression")
                        if compression == "gzip" or compression == "zlib":
                            try:
                                import zlib
                                # CTest writes a zlib deflate stream under compression='gzip'
                                decoded_data = zlib.decompress(decoded_data)
                            except Exception:
                                try:
                                    import gzip
                                    decoded_data = gzip.decompress(decoded_data)
                                except Exception:
                                    pass
                        output = decoded_data.decode("utf-8", errors="ignore")
                    except Exception as e:
                        print(f"Error decoding log: {e}")
                        output = raw_text
                else:
                    output = raw_text

        tests.append({
            "name": name,
            "status": status.capitalize(),
            "duration": float(duration) if duration else 0.0,
            "output": output.strip()
        })
    return tests

def generate_excel_report(output_file):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    # Find all Test.xml files recursively
    xml_files = glob.glob("**/Testing/*/Test.xml", recursive=True)
    if not xml_files:
        print("No CTest Test.xml files found in the workspace.")
        return False

    if has_openpyxl:
        print("Using openpyxl to generate native Excel (.xlsx) file...")
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for xml_path in xml_files:
            # Determine sheet name based on the parent folder (e.g. build-linux, build-sanitizers)
            parts = xml_path.replace("\\", "/").split("/")
            sheet_name = "Tests"
            for part in parts:
                if part.startswith("build-"):
                    sheet_name = part.replace("build-", "").capitalize()
                    break
            
            # Excel sheet names have a 31 char limit
            sheet_name = sheet_name[:30]
            ws = wb.create_sheet(title=sheet_name)
            
            # Setup sheet gridlines visible
            ws.views.sheetView[0].showGridLines = True
            
            # Get test list
            tests = parse_ctest_xml(xml_path)
            
            # Calculate summary stats
            total_tests = len(tests)
            passed_tests = sum(1 for t in tests if t["status"].lower() == "passed")
            failed_tests = total_tests - passed_tests
            success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            # Colors and Fills
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark blue
            sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Soft blue tint
            pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
            fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
            
            font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Calibri", size=11, bold=True)
            font_regular = Font(name="Calibri", size=11)
            font_pass = Font(name="Calibri", size=11, color="375623", bold=True)
            font_fail = Font(name="Calibri", size=11, color="C00000", bold=True)
            
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF')
            )
            
            # Write Title and Summary block
            ws.cell(row=1, column=1, value=f"{sheet_name} Test Execution Report").font = font_title
            
            ws.cell(row=3, column=1, value="Metric").font = font_bold
            ws.cell(row=3, column=1).fill = sub_header_fill
            ws.cell(row=3, column=1).border = thin_border
            
            ws.cell(row=3, column=2, value="Value").font = font_bold
            ws.cell(row=3, column=2).fill = sub_header_fill
            ws.cell(row=3, column=2).border = thin_border
            
            metrics = [
                ("Total Tests Run", total_tests),
                ("Passed Tests", passed_tests),
                ("Failed Tests", failed_tests),
                ("Success Rate", f"{success_rate:.1f}%")
            ]
            
            for i, (m, v) in enumerate(metrics, start=4):
                ws.cell(row=i, column=1, value=m).font = font_regular
                ws.cell(row=i, column=1).border = thin_border
                
                val_cell = ws.cell(row=i, column=2, value=v)
                val_cell.font = font_bold
                val_cell.border = thin_border
                if m == "Passed Tests" and passed_tests > 0:
                    val_cell.fill = pass_fill
                    val_cell.font = font_pass
                elif m == "Failed Tests" and failed_tests > 0:
                    val_cell.fill = fail_fill
                    val_cell.font = font_fail
            
            # Write Table Headers
            headers = ["Test Name", "Status", "Duration (sec)", "Output Logs"]
            start_row = 10
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            
            # Write Test rows
            for row_idx, test in enumerate(tests, start=start_row + 1):
                c1 = ws.cell(row=row_idx, column=1, value=test["name"])
                c2 = ws.cell(row=row_idx, column=2, value=test["status"])
                c3 = ws.cell(row=row_idx, column=3, value=test["duration"])
                c4 = ws.cell(row=row_idx, column=4, value=test["output"][:30000]) # Cap output text in cell
                
                for c in [c1, c2, c3, c4]:
                    c.font = font_regular
                    c.border = thin_border
                
                c3.alignment = Alignment(horizontal="right")
                c4.alignment = Alignment(wrap_text=True)
                
                # Highlight status
                if test["status"].lower() == "passed":
                    c2.fill = pass_fill
                    c2.font = font_pass
                    c2.alignment = Alignment(horizontal="center")
                else:
                    c2.fill = fail_fill
                    c2.font = font_fail
                    c2.alignment = Alignment(horizontal="center")
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < start_row:
                        continue # Skip title/summary for auto-width
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            # Give column 4 (logs) a wider default width
            ws.column_dimensions['D'].width = 50

        wb.save(output_file)
        print(f"Excel report generated successfully at: {output_file}")
        return True
    else:
        print("openpyxl is not installed. Generating HTML-based Excel report instead...")
        # Write HTML table report which Excel opens natively
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("""<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel" xmlns="http://www.w3.org/TR/REC-html40">
<head>
<meta http-equiv="content-type" content="text/html; charset=utf-8">
<style>
  body { font-family: Calibri, sans-serif; }
  table { border-collapse: collapse; margin-bottom: 30px; }
  th { background-color: #1F497D; color: white; font-weight: bold; }
  td, th { border: 1px solid #BFBFBF; text-align: left; padding: 8px; }
  .metric-header { background-color: #DCE6F1; font-weight: bold; }
  .passed { background-color: #E2EFDA; color: #375623; font-weight: bold; text-align: center; }
  .failed { background-color: #FCE4D6; color: #C00000; font-weight: bold; text-align: center; }
  .title { font-size: 16px; font-weight: bold; color: #1F497D; margin-bottom: 10px; }
</style>
</head>
<body>
""")
            for xml_path in xml_files:
                parts = xml_path.replace("\\", "/").split("/")
                title = "Tests"
                for part in parts:
                    if part.startswith("build-"):
                        title = part.replace("build-", "").capitalize()
                        break
                
                tests = parse_ctest_xml(xml_path)
                total_tests = len(tests)
                passed_tests = sum(1 for t in tests if t["status"].lower() == "passed")
                failed_tests = total_tests - passed_tests
                success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
                
                f.write(f'<div class="title">{title} Test Execution Report</div>\n')
                f.write('<table>\n')
                f.write(f'  <tr><td class="metric-header">Total Tests Run</td><td><b>{total_tests}</b></td></tr>\n')
                f.write(f'  <tr><td class="metric-header">Passed Tests</td><td class="{ "passed" if passed_tests > 0 else "" }">{passed_tests}</td></tr>\n')
                f.write(f'  <tr><td class="metric-header">Failed Tests</td><td class="{ "failed" if failed_tests > 0 else "" }">{failed_tests}</td></tr>\n')
                f.write(f'  <tr><td class="metric-header">Success Rate</td><td><b>{success_rate:.1f}%</b></td></tr>\n')
                f.write('</table>\n')
                
                f.write('<table>\n')
                f.write('  <tr><th>Test Name</th><th>Status</th><th>Duration (sec)</th><th>Output Logs</th></tr>\n')
                for test in tests:
                    status_class = "passed" if test["status"].lower() == "passed" else "failed"
                    # Escape HTML characters in output logs
                    safe_output = test["output"].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    f.write(f'  <tr><td>{test["name"]}</td><td class="{status_class}">{test["status"]}</td><td style="text-align: right;">{test["duration"]:.3f}</td><td><pre style="margin:0;">{safe_output[:30000]}</pre></td></tr>\n')
                f.write('</table><br>\n')
            f.write("</body>\n</html>\n")
        print(f"HTML Excel report generated successfully at: {output_file}")
        return True

if __name__ == "__main__":
    out_file = "test_report.xls"
    if len(sys.argv) > 1:
        out_file = sys.argv[1]
    
    # Ensure correct extension based on package availability
    try:
        import openpyxl
        # If openpyxl is present, write standard .xlsx file.
        # But if the user explicitly requested .xls, name it accordingly.
        if out_file.endswith(".xls") and not out_file.endswith(".xlsx"):
            # openpyxl only writes xlsx, so rename or let it write xlsx
            pass
    except ImportError:
        pass
        
    generate_excel_report(out_file)
